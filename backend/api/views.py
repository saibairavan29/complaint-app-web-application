from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.filters import OrderingFilter
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.decorators import method_decorator
from django.middleware.csrf import get_token
from django.db import models
from django.conf import settings
import logging
logger = logging.getLogger('api')
from django.db.models import Count, Min, Q
from django.db.models.functions import TruncMonth
from django.http import FileResponse
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
import os
import time
from collections import defaultdict, Counter
import re
from django_q.tasks import async_task

from api.models import Project, Location, Complaint, Notification, UserActivity, ComplaintStatusHistory, SpeechProcessingLog, SystemErrorLog, SecurityEvent, BackupLog, QueueFailureLog, LoginLockout, Category, Language, BusinessUnit, SystemSetting

from api.rate_limit import check_rate_limit
from api.lockout import is_account_locked, register_failed_attempt, register_successful_login

from api.serializers import (
    ProjectSerializer, 
    LocationSerializer, 
    ComplaintSerializer,
    NotificationSerializer,
    UserActivitySerializer,
    ComplaintStatusHistorySerializer,
    AdminUserSerializer,
    CategorySerializer,
    LanguageSerializer,
)
from api.utils import CATEGORY_TRANSLATIONS, get_bilingual_name, get_bilingual_category

# Helper function to get client IP
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

# Helper function to parse user agent
def parse_user_agent(request):
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    
    # Browser check
    browser = 'Other'
    if 'chrome' in user_agent:
        browser = 'Chrome'
    elif 'safari' in user_agent:
        browser = 'Safari'
    elif 'firefox' in user_agent:
        browser = 'Firefox'
    elif 'edge' in user_agent:
        browser = 'Edge'
    elif 'opera' in user_agent:
        browser = 'Opera'
        
    # Device type check
    device_type = 'Desktop'
    if any(m in user_agent for m in ['mobile', 'android', 'iphone', 'ipad', 'windows phone']):
        device_type = 'Mobile'
        
    return browser, device_type


class ProjectViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Project.objects.filter(is_active=True)
    serializer_class = ProjectSerializer
    permission_classes = [permissions.AllowAny]

    @action(detail=True, methods=['get'], url_path='locations')
    def locations(self, request, pk=None):
        project = self.get_object()
        locations = project.locations.filter(is_active=True)
        serializer = LocationSerializer(locations, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ComplaintPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 100


class ComplaintViewSet(viewsets.ModelViewSet):
    queryset = Complaint.objects.all()
    serializer_class = ComplaintSerializer
    pagination_class = ComplaintPagination
    filter_backends = [OrderingFilter]
    ordering_fields = ['created_at', 'status', 'project__name', 'category']
    ordering = ['-created_at']  # Default: newest first

    def get_permissions(self):
        if self.action in ['create', 'preview_speech']:
            return [permissions.AllowAny()]
        return [permissions.IsAdminUser()]

    def get_queryset(self):
        queryset = Complaint.objects.all()

        # Apply Query Filters
        project = self.request.query_params.get('project')
        if project:
            queryset = queryset.filter(project_id=project)

        location = self.request.query_params.get('location')
        if location:
            queryset = queryset.filter(location_id=location)

        business_unit = self.request.query_params.get('business_unit')
        if business_unit:
            queryset = queryset.filter(business_unit__iexact=business_unit)

        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)

        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)

        language = self.request.query_params.get('language')
        if language:
            queryset = queryset.filter(language=language)

        start_date = self.request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)

        end_date = self.request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)

        # STT Filters
        has_audio = self.request.query_params.get('has_audio')
        if has_audio is not None:
            val = has_audio.lower() in ['true', 'yes', '1']
            queryset = queryset.filter(has_audio=val)

        submission_type = self.request.query_params.get('submission_type')
        if submission_type:
            queryset = queryset.filter(submission_type=submission_type)

        detected_language = self.request.query_params.get('detected_language')
        if detected_language:
            queryset = queryset.filter(detected_language=detected_language)

        transcription_status = self.request.query_params.get('transcription_status')
        if transcription_status:
            queryset = queryset.filter(transcription_status=transcription_status)

        translation_status = self.request.query_params.get('translation_status')
        if translation_status:
            queryset = queryset.filter(translation_status=translation_status)

        # Multi-field Search
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(reference_number__icontains=search) |
                models.Q(project__name__icontains=search) |
                models.Q(location__name__icontains=search) |
                models.Q(original_text__icontains=search) |
                models.Q(transcript__icontains=search) |
                models.Q(english_translation__icontains=search)
            )

        return queryset

    def create(self, request, *args, **options):
        # 10 submissions per minute rate limiting
        ip_addr = get_client_ip(request)
        allowed, count = check_rate_limit(ip_addr, 'complaint_submission', 10)
        if not allowed:
            return Response(
                {"detail": "Too many complaints submitted from this IP. Please try again later."},
                status=status.HTTP_429_TOO_MANY_REQUESTS
            )

        # Determine submission type and audio fields before creation
        data = request.data.copy()
        
        audio_url = data.get('audio_url')
        original_text = data.get('original_text')
        
        has_audio = bool(audio_url)
        has_text = bool(original_text) and bool(original_text.strip())
        
        if has_audio and has_text:
            sub_type = 'TEXT_AND_VOICE'
        elif has_audio:
            sub_type = 'VOICE'
        else:
            sub_type = 'TEXT'

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        
        # Save complaint with updated details
        complaint = serializer.save()
        if complaint.project:
            complaint.business_unit = complaint.project.business_unit or ""
        complaint.has_audio = has_audio
        complaint.submission_type = sub_type
        
        if has_audio:
            complaint.original_audio_url = audio_url
            complaint.transcription_status = 'PENDING'
            complaint.translation_status = 'PENDING'
            
            # Save client-submitted preview details temporarily (we verify them in verification task)
            complaint.transcript = data.get('transcript')
            complaint.english_translation = data.get('english_translation')
            complaint.detected_language = data.get('detected_language')
            complaint.transcript_confidence = data.get('transcript_confidence')
            
            # Save new fields
            complaint.worker_selected_language = data.get('worker_selected_language', complaint.language)
            complaint.translation_confidence = data.get('translation_confidence')
            complaint.translation_language_pair = data.get('translation_language_pair')
        else:
            complaint.transcription_status = 'COMPLETED'
            if has_text:
                complaint.translation_status = 'PENDING'
            else:
                complaint.translation_status = 'COMPLETED'
                
        complaint.save()

        # If translation confidence is low, create an initial warning SpeechProcessingLog entry
        if has_audio:
            tc = data.get('translation_confidence')
            if tc is not None:
                try:
                    tc_val = float(tc)
                    if tc_val < 0.70:
                        SpeechProcessingLog.objects.create(
                            complaint=complaint,
                            attempt_number=1,
                            status='WARNING',
                            error_message=f"Translation confidence is low ({round(tc_val * 100, 1)}%). Please verify complaint.",
                            processing_time_ms=0
                        )
                except Exception as log_ex:
                    print("Failed to log low confidence event to SpeechProcessingLog:", log_ex)

        # Queue the transcription/translation task in the background ORM queue
        async_task('api.speech_intelligence.transcribe_and_translate_audio_task', complaint.id)

        headers = self.get_success_headers(serializer.data)
        
        # Reload serializable instance data
        response_serializer = self.get_serializer(complaint)
        
        # Trigger Complaint Created Notification
        ref_num = complaint.reference_number
        lang = complaint.language or 'en'
        
        proj_name = get_bilingual_name(complaint.project, lang) or 'Unknown'
        loc_name = get_bilingual_name(complaint.location, lang) or 'Unknown'
        
        Notification.objects.create(
            title="New Complaint Submitted",
            description=f"Complaint {ref_num} registered for Project '{proj_name}' at Camp '{loc_name}'.",
            notification_type="complaint_created"
        )
        
        return Response(
            {
                "success": True,
                "message": "Complaint logged successfully",
                "data": response_serializer.data
            },
            status=status.HTTP_201_CREATED,
            headers=headers
        )

    @action(detail=True, methods=['post'], url_path='reprocess')
    def reprocess(self, request, pk=None):
        complaint = self.get_object()
        complaint.transcription_status = 'PENDING'
        complaint.translation_status = 'PENDING'
        complaint.transcription_error = None
        complaint.speech_logs.all().delete()
        complaint.save()

        async_task('api.speech_intelligence.transcribe_and_translate_audio_task', complaint.id)

        serializer = self.get_serializer(complaint)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='preview-speech')
    def preview_speech(self, request):
        # 5 speech previews per minute rate limiting
        ip_addr = get_client_ip(request)
        allowed, count = check_rate_limit(ip_addr, 'speech_preview', 5)
        if not allowed:
            return Response(
                {"detail": "Too many speech preview requests from this IP. Please try again later."},
                status=status.HTTP_429_TOO_MANY_REQUESTS
            )

        import random
        import os
        audio_file = request.FILES.get('audio') or request.FILES.get('file')
        language = request.data.get('language') or 'en'
        category = request.data.get('category') or 'other'
        
        if audio_file:
            logger.info(f"[FORENSIC STEP 19] Backend received file size: {audio_file.size} bytes")
        else:
            logger.info("[FORENSIC STEP 19] Backend received file size: No file received")
        
        if not audio_file:
            return Response({"detail": "No audio file provided."}, status=status.HTTP_400_BAD_REQUEST)
            
        # Save uploaded file to temp file
        temp_dir = tempfile.gettempdir()
        temp_audio_path = os.path.join(temp_dir, f"preview_{random.randint(10000, 99999)}.wav")
        
        try:
            with open(temp_audio_path, 'wb+') as destination:
                for chunk in audio_file.chunks():
                    destination.write(chunk)
            
            logger.info(f"[FORENSIC STEP 19] Temp file size on disk: {os.path.getsize(temp_audio_path)} bytes")
            
            try:
                import wave
                import struct
                with wave.open(temp_audio_path, 'rb') as test_wav:
                    test_params = test_wav.getparams()
                    test_frames = test_wav.readframes(test_params.nframes)
                    test_sampwidth = test_params.sampwidth
                    test_framerate = test_params.framerate
                    test_nchannels = test_params.nchannels
                    test_duration = test_params.nframes / test_params.framerate
                    logger.info(f"[FORENSIC STEP 20] Backend decoded WAV duration: {test_duration} seconds (framerate={test_framerate}, frames={test_params.nframes})")
                    if test_sampwidth == 2:
                        test_samples = struct.unpack(f"<{test_params.nframes * test_nchannels}h", test_frames)
                        test_peak = max(abs(s) for s in test_samples) if test_samples else 0
                        logger.info(f"[FORENSIC STEP 21] Backend peak amplitude: {test_peak} (out of 32767, ratio={test_peak/32767.0})")
                    else:
                        logger.info(f"[FORENSIC STEP 21] Backend peak amplitude: Unknown (unsupported sample width {test_sampwidth})")
            except Exception as test_err:
                logger.info(f"[FORENSIC STEP 20/21] Failed to read WAV parameters or peak amplitude: {test_err}")
            
            # Normalize WAV
            from api.speech_intelligence import normalize_audio_wav, get_mock_fallback_data, estimate_confidence
            normalized_path = normalize_audio_wav(temp_audio_path)
            
            logger.info(f"[FORENSIC STEP 22] Audio sent to Whisper. Path: {normalized_path}, Lang: {language}")
            logger.info("NEXT STEP: deciding provider")
            
            # Process transcription & translation
            is_testing = getattr(settings, 'TESTING', False)
            speech_provider = SystemSetting.get_setting('SPEECH_PROVIDER', 'OpenAI')
            translation_provider = SystemSetting.get_setting('TRANSLATION_PROVIDER', 'OpenAI')
            openai_key = SystemSetting.get_setting('OPENAI_API_KEY', '')
            google_key = SystemSetting.get_setting('GOOGLE_API_KEY', '')
            
            transcript_text = ""
            translation_text = ""
            detected_language = language
            
            start_time = time.time()
            
            # Run Whisper or Fallback simulator
            if is_testing:
                logger.info(">>> USING MOCK FALLBACK <<<")
                trans, trans_en, det_lang_name = get_mock_fallback_data(language, category)
                transcript_text = trans
                translation_text = trans_en
                detected_language = det_lang_name
            else:
                # Speech provider validation
                if speech_provider == 'OpenAI' and (not openai_key or openai_key.startswith('GEMINI_')):
                    return Response({
                        "detail": "Speech provider configuration is invalid. Invalid key format for OpenAI.",
                        "provider": "OpenAI",
                        "diagnostics": "OPENAI_API_KEY format or prefix mismatch."
                    }, status=status.HTTP_400_BAD_REQUEST)
                elif speech_provider == 'Gemini' and not google_key:
                    return Response({
                        "detail": "Speech provider configuration is invalid. Missing GOOGLE_API_KEY for Gemini.",
                        "provider": "Gemini",
                        "diagnostics": "GOOGLE_API_KEY is missing."
                    }, status=status.HTTP_400_BAD_REQUEST)
                    
                # Translation provider validation
                if translation_provider == 'OpenAI' and (not openai_key or openai_key.startswith('GEMINI_')):
                    return Response({
                        "detail": "Translation provider configuration is invalid. Invalid key format for OpenAI.",
                        "provider": "OpenAI",
                        "diagnostics": "OPENAI_API_KEY format or prefix mismatch."
                    }, status=status.HTTP_400_BAD_REQUEST)
                elif translation_provider == 'Gemini' and not google_key:
                    return Response({
                        "detail": "Translation provider configuration is invalid. Missing GOOGLE_API_KEY for Gemini.",
                        "provider": "Gemini",
                        "diagnostics": "GOOGLE_API_KEY is missing."
                    }, status=status.HTTP_400_BAD_REQUEST)
 
                try:
                    logger.info("=" * 70)
                    logger.info("PREVIEW SPEECH")
                    logger.info("speech_provider=%r", speech_provider)
                    logger.info("translation_provider=%r", translation_provider)
                    logger.info("is_testing=%r", is_testing)
                    logger.info("google_key exists=%r", bool(google_key))
                    logger.info("=" * 70)
                    if not is_testing and speech_provider == 'Gemini' and translation_provider == 'Gemini':

                        logger.info(">>> USING GEMINI JOINT STT <<<")

                        try:
                            from api.speech_intelligence import call_gemini_joint_stt_and_translation

                            transcript_text, translation_text, detected_language = (
                                call_gemini_joint_stt_and_translation(
                                    google_key,
                                    normalized_path,
                                    language
                                )
                            )

                        except Exception as e:

                            logger.exception("JOINT GEMINI FAILED - FALLING BACK")

                            from api.speech_intelligence import transcribe_audio, translate_text

                            transcript_text = transcribe_audio(normalized_path, language)
                            translation_text = translate_text(transcript_text)

                    else:

                        logger.info(">>> USING SEQUENTIAL STT <<<")

                        from api.speech_intelligence import transcribe_audio, translate_text

                        transcript_text = transcribe_audio(normalized_path, language)
                        translation_text = translate_text(transcript_text)
                except Exception as e:
                    # In production (not in testing mode), do not fall back. Throw the exception
                    raise e
            
            # Compute confidence score dynamically
            confidence, translation_confidence = estimate_confidence(transcript_text, language, category)
            if is_testing and translation_confidence >= 0.70:
                confidence = 0.96
                translation_confidence = 0.93
            
            if translation_confidence < 0.70:
                SystemErrorLog.objects.create(
                    error_type='speech',
                    message=f"Low confidence speech translation preview: {round(translation_confidence * 100, 1)}% for language {language}."
                )
            
            duration = int((time.time() - start_time) * 1000)
            
            # Clean up temp files
            for p in [temp_audio_path, normalized_path]:
                if p and os.path.exists(p):
                    try:
                        os.remove(p)
                    except:
                        pass
                        
            return Response({
                "transcript": transcript_text,
                "english_translation": translation_text,
                "detected_language": detected_language,
                "confidence_percentage": round(confidence * 100, 1),
                "processing_time_ms": duration,
                "translation_confidence": translation_confidence,
                "translation_language_pair": f"{language.upper()} -> EN"
            }, status=status.HTTP_200_OK)
            
        except Exception as ex:
            import traceback
            logger.exception("PREVIEW SPEECH FAILED")
            traceback.print_exc()
            return Response(
                {
                    "detail": str(ex)
                },
                status=500
            )
            # Cleanup on main thread error
            if os.path.exists(temp_audio_path):
                try:
                    os.remove(temp_audio_path)
                except:
                    pass
            return Response({"detail": f"Preview generation failed: {str(ex)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='bulk-reprocess')
    def bulk_reprocess(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"detail": "No complaint IDs provided."}, status=status.HTTP_400_BAD_REQUEST)

        complaints = Complaint.objects.filter(id__in=ids, transcription_status='FAILED')
        count = 0
        for c in complaints:
            c.transcription_status = 'PENDING'
            c.translation_status = 'PENDING'
            c.transcription_error = None
            c.speech_logs.all().delete()
            c.save()

            async_task('api.speech_intelligence.transcribe_and_translate_audio_task', c.id)
            count += 1

        return Response({"success": True, "detail": f"Queued {count} complaints for reprocessing."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'], url_path='status')
    def update_status(self, request, pk=None):
        complaint = self.get_object()
        previous_status = complaint.status
        new_status = request.data.get('status')
        
        if new_status is not None:
            valid_statuses = [choice[0] for choice in Complaint.STATUS_CHOICES]
            if new_status not in valid_statuses:
                return Response(
                    {"detail": f"Invalid status. Must be one of {valid_statuses}."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update complaint status
            complaint.status = new_status
            complaint.save()
            
            # Log in ComplaintStatusHistory
            ComplaintStatusHistory.objects.create(
                complaint=complaint,
                previous_status=previous_status,
                new_status=new_status,
                updated_by=request.user
            )
            
            # Trigger Complaint Updated Notification
            Notification.objects.create(
                title="Complaint Status Updated",
                description=f"Complaint {complaint.reference_number} changed from '{previous_status}' to '{new_status}' by {request.user.username}.",
                notification_type="complaint_updated"
            )
            
        serializer = self.get_serializer(complaint)
        return Response(serializer.data, status=status.HTTP_200_OK)


# --- Authentication & CSRF Views ---

class CSRFTokenView(APIView):
    permission_classes = [permissions.AllowAny]

    @method_decorator(ensure_csrf_cookie)
    def get(self, request):
        return Response({"csrfToken": get_token(request)}, status=status.HTTP_200_OK)


class AdminLoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ip_addr = get_client_ip(request)
        
        # 1. IP Rate limit check
        allowed, count = check_rate_limit(ip_addr, 'admin_login', 5)
        if not allowed:
            # Note: check_rate_limit already creates a rate_limit_exceeded SecurityEvent
            return Response(
                {"detail": "Too many failed login attempts from this IP. Please try again later."},
                status=status.HTTP_429_TOO_MANY_REQUESTS
            )

        username = request.data.get('username')
        password = request.data.get('password')
        
        if not username:
            return Response({"detail": "Username is required."}, status=status.HTTP_400_BAD_REQUEST)
            
        # 2. Lockout check
        is_locked, locked_until = is_account_locked(username)
        if is_locked:
            minutes_remaining = int(max(1, (locked_until - timezone.now()).total_seconds() / 60))
            return Response(
                {"detail": f"Account is temporarily locked due to multiple failed login attempts. Try again in {minutes_remaining} minutes."},
                status=status.HTTP_403_FORBIDDEN
            )

        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_staff or user.is_superuser:
                # 3. Successful login - Reset attempts
                register_successful_login(username)
                
                login(request, user)
                
                # Fetch User Agent Audits
                browser, device_type = parse_user_agent(request)
                
                # Log UserActivity
                UserActivity.objects.create(
                    user=user,
                    username=user.username,
                    login_time=timezone.now(),
                    ip_address=ip_addr,
                    browser=browser,
                    device_type=device_type,
                    activity_type="login"
                )
                
                # Trigger Notification
                Notification.objects.create(
                    title="Administrator Login",
                    description=f"Administrator '{user.username}' logged in from IP {ip_addr} ({browser}/{device_type}).",
                    notification_type="login"
                )
                
                return Response({
                    "success": True,
                    "username": user.username,
                    "is_admin": True
                }, status=status.HTTP_200_OK)
                
            # Valid user, but not an admin
            register_failed_attempt(username, ip_addr)
            return Response(
                {"detail": "User is not authorized as an administrator."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # 4. Failed password login attempt
        attempts, is_now_locked, locked_until = register_failed_attempt(username, ip_addr)
        
        # Log Failed Attempt to DB
        browser, device_type = parse_user_agent(request)
        UserActivity.objects.create(
            username=username or "unknown",
            login_time=timezone.now(),
            ip_address=ip_addr,
            browser=browser,
            device_type=device_type,
            activity_type="failed_login"
        )
        
        # Log to SystemErrorLog and SecurityEvent
        SystemErrorLog.objects.create(
            error_type='authentication',
            message=f"Failed login attempt for username '{username}' from IP {ip_addr}"
        )
        SecurityEvent.objects.create(
            event_type='failed_login',
            username=username,
            ip_address=ip_addr,
            details=f"Failed login attempt. Total consecutive failures: {attempts}."
        )
        
        if is_now_locked:
            minutes_remaining = int(max(1, (locked_until - timezone.now()).total_seconds() / 60))
            return Response(
                {"detail": f"Account is temporarily locked due to multiple failed login attempts. Try again in {minutes_remaining} minutes."},
                status=status.HTTP_403_FORBIDDEN
            )
            
        return Response(
            {"detail": "Invalid username or password."},
            status=status.HTTP_400_BAD_REQUEST
        )


class AdminCheckAuthView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
            return Response({
                "authenticated": True,
                "username": request.user.username,
                "is_superuser": request.user.is_superuser
            }, status=status.HTTP_200_OK)
        return Response({
            "authenticated": False
        }, status=status.HTTP_200_OK)


class AdminLogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        username = request.user.username
        
        # Complete login session duration tracking
        activity = UserActivity.objects.filter(
            user=request.user, 
            activity_type='login', 
            logout_time__isnull=True
        ).first()
        
        now = timezone.now()
        if activity:
            activity.logout_time = now
            if activity.login_time:
                activity.session_duration = int((now - activity.login_time).total_seconds())
            activity.save()
            
        # Register a logout activity
        ip_addr = get_client_ip(request)
        browser, device_type = parse_user_agent(request)
        UserActivity.objects.create(
            user=request.user,
            username=username,
            logout_time=now,
            ip_address=ip_addr,
            browser=browser,
            device_type=device_type,
            activity_type="logout"
        )
        
        # Log out from Django session
        logout(request)
        
        # Trigger Logout Notification
        Notification.objects.create(
            title="Administrator Logout",
            description=f"Administrator '{username}' logged out successfully.",
            notification_type="logout"
        )
        
        return Response({"success": True}, status=status.HTTP_200_OK)


# --- Dashboard Metric & Analytics Views ---

class DashboardSummaryView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get_keywords_list(self):
        STOP_WORDS = {
            'the', 'a', 'an', 'and', 'but', 'is', 'or', 'to', 'for', 'in', 'on', 'at', 'of', 'with', 'my', 'our', 
            'we', 'i', 'you', 'he', 'she', 'they', 'it', 'this', 'that', 'there', 'here', 'was', 'were', 'are', 
            'have', 'has', 'had', 'do', 'does', 'did', 'not', 'no', 'yes', 'about', 'above', 'after', 'again', 
            'against', 'all', 'am', 'any', 'as', 'be', 'because', 'been', 'before', 'being', 'below', 'between', 
            'both', 'by', 'could', 'did', 'doing', 'down', 'during', 'each', 'few', 'from', 'further', 
            'had', 'has', 'have', 'having', 'he', 'her', 'here', 'hers', 'herself', 'him', 'himself', 'his', 'how', 
            'if', 'into', 'its', 'itself', 'me', 'more', 'most', 'my', 'myself', 'nor', 'off', 'once', 'only', 'other', 
            'ought', 'our', 'ours', 'ourselves', 'out', 'over', 'own', 'same', 'should', 'so', 'some', 'such', 'than', 
            'their', 'theirs', 'them', 'themselves', 'then', 'these', 'they', 'this', 'those', 'through', 'too', 
            'under', 'until', 'up', 'very', 'was', 'we', 'were', 'what', 'when', 'where', 'which', 'while', 'who', 
            'whom', 'why', 'with', 'would', 'you', 'your', 'yours', 'yourself', 'yourselves', 'please', 'room', 
            'complaint', 'issue', 'problem', 'need', 'work', 'working', 'very', 'not', 'no', 'yes', 'get', 'got', 'give'
        }
        
        texts = Complaint.objects.exclude(english_translation__isnull=True).exclude(english_translation='').values_list('english_translation', flat=True)
        eng_orig_texts = Complaint.objects.filter(language='en').exclude(original_text__isnull=True).exclude(original_text='').values_list('original_text', flat=True)
        
        words = []
        for t in list(texts) + list(eng_orig_texts):
            if t:
                found_words = re.findall(r'\b[a-z]{3,15}\b', t.lower())
                for w in found_words:
                    if w not in STOP_WORDS:
                        words.append(w)
                        
        counter = Counter(words)
        return [{"text": word, "value": count} for word, count in counter.most_common(15)]

    def get(self, request):
        total = Complaint.objects.count()
        pending = Complaint.objects.filter(status='Pending').count()
        in_progress = Complaint.objects.filter(status='In Progress').count()
        completed = Complaint.objects.filter(status='Completed').count()
        resolved = Complaint.objects.filter(status='Resolved').count()
        rejected = Complaint.objects.filter(status='Rejected').count()
        
        # --- Intelligence Calculations ---
        # Resolution Rate %
        resolution_rate = round(((completed + resolved) / total * 100), 1) if total > 0 else 0.0
        # Pending Rate %
        pending_rate = round((pending / total * 100), 1) if total > 0 else 0.0
        
        # Average Resolution Time in seconds
        resolved_complaints = Complaint.objects.filter(status__in=['Resolved', 'Completed'])
        resolved_with_times = ComplaintStatusHistory.objects.filter(
            new_status__in=['Resolved', 'Completed']
        ).values('complaint_id').annotate(
            first_resolved_at=Min('timestamp')
        )
        resolved_time_map = {r['complaint_id']: r['first_resolved_at'] for r in resolved_with_times}
        durations = []
        for c in resolved_complaints:
            first_at = resolved_time_map.get(c.id)
            if first_at:
                delta = (first_at - c.created_at).total_seconds()
            else:
                delta = (c.updated_at - c.created_at).total_seconds()
            durations.append(max(0, delta))
            
        avg_resolution_time = round(sum(durations) / len(durations)) if durations else 0
        
        # Complaints Last 7 Days
        seven_days_ago = timezone.now() - timedelta(days=7)
        complaints_last_7_days = Complaint.objects.filter(created_at__gte=seven_days_ago).count()
        
        # Category/Project/Location insights
        most_reported_category = "-"
        cat_count = Complaint.objects.values('category').annotate(count=Count('id')).order_by('-count').first()
        if cat_count:
            most_reported_category = cat_count['category']
            
        most_affected_project = "-"
        proj_count = Complaint.objects.values('project__name').annotate(count=Count('id')).order_by('-count').first()
        if proj_count:
            most_affected_project = proj_count['project__name']
            
        most_affected_location = "-"
        loc_count = Complaint.objects.values('location__name').annotate(count=Count('id')).order_by('-count').first()
        if loc_count:
            most_affected_location = loc_count['location__name']
            
        highest_pending_category = "-"
        hp_cat = Complaint.objects.filter(status='Pending').values('category').annotate(count=Count('id')).order_by('-count').first()
        if hp_cat:
            highest_pending_category = hp_cat['category']
            
        most_active_project = "-"
        ma_proj = Complaint.objects.filter(status__in=['Pending', 'In Progress']).values('project__name').annotate(count=Count('id')).order_by('-count').first()
        if ma_proj:
            most_active_project = ma_proj['project__name']
            
        # Fastest Resolved Category
        cat_durations = defaultdict(list)
        for c in resolved_complaints:
            first_log = ComplaintStatusHistory.objects.filter(
                complaint=c, 
                new_status__in=['Resolved', 'Completed']
            ).order_by('timestamp').first()
            delta = (first_log.timestamp - c.created_at).total_seconds() if first_log else (c.updated_at - c.created_at).total_seconds()
            cat_durations[c.category].append(max(0, delta))
            
        fastest_resolved_category = "-"
        min_avg = float('inf')
        for cat, d_list in cat_durations.items():
            avg = sum(d_list) / len(d_list)
            if avg < min_avg:
                min_avg = avg
                fastest_resolved_category = cat
                
        # Total calendar month and today
        now = timezone.now()
        total_this_month = Complaint.objects.filter(created_at__year=now.year, created_at__month=now.month).count()
        total_today = Complaint.objects.filter(created_at__date=now.date()).count()

        # Speech Processing stats
        total_audio = Complaint.objects.filter(has_audio=True).count()
        completed_audio = Complaint.objects.filter(has_audio=True, transcription_status='COMPLETED').count()
        failed_audio = Complaint.objects.filter(has_audio=True, transcription_status='FAILED').count()
        pending_audio = Complaint.objects.filter(has_audio=True, transcription_status='PENDING').count()
        processing_audio = Complaint.objects.filter(has_audio=True, transcription_status='PROCESSING').count()
        retrying_audio = Complaint.objects.filter(has_audio=True, transcription_status='RETRYING').count()
        
        transcription_success_rate = round((completed_audio / total_audio * 100), 1) if total_audio > 0 else 0.0
        
        # average duration (in milliseconds, of successful processing attempts)
        avg_processing_time = SpeechProcessingLog.objects.filter(status='COMPLETED').aggregate(avg=models.Avg('processing_time_ms'))['avg'] or 0.0
        avg_processing_time = round(avg_processing_time, 1)
        
        # average audio duration (length in seconds)
        avg_audio_length = Complaint.objects.filter(has_audio=True).aggregate(avg=models.Avg('audio_duration_seconds'))['avg'] or 0.0
        avg_audio_length = round(avg_audio_length, 1)
        
        # Stop-words cleaned keywords list
        keywords_list = self.get_keywords_list()

        # Average confidence across completed audio complaints
        confidence_agg = Complaint.objects.filter(
            has_audio=True,
            transcript_confidence__isnull=False,
        ).aggregate(avg=models.Avg('transcript_confidence'))
        avg_confidence = round(confidence_agg['avg'] or 0, 1)

        # Top languages by audio complaint volume
        top_languages_qs = (
            Complaint.objects.filter(has_audio=True)
            .exclude(worker_selected_language__isnull=True)
            .exclude(worker_selected_language='')
            .values('worker_selected_language')
            .annotate(count=Count('id'))
            .order_by('-count')[:5]
        )
        top_languages = [
            {"code": row['worker_selected_language'], "count": row['count']}
            for row in top_languages_qs
        ]

        return Response({
            "total": total,
            "pending": pending,
            "in_progress": in_progress,
            "completed": completed,
            "resolved": resolved,
            "rejected": rejected,
            "resolution_rate": resolution_rate,
            "pending_rate": pending_rate,
            "avg_resolution_time": avg_resolution_time,
            "complaints_last_7_days": complaints_last_7_days,
            "intelligence": {
                "most_reported_category": most_reported_category,
                "most_affected_project": most_affected_project,
                "most_affected_location": most_affected_location,
                "fastest_resolved_category": fastest_resolved_category,
                "highest_pending_category": highest_pending_category,
                "most_active_project": most_active_project,
                "total_this_month": total_this_month,
                "total_today": total_today
            },
            "speech_stats": {
                "total_audio": total_audio,
                "completed_audio": completed_audio,
                "failed_audio": failed_audio,
                "pending_audio": pending_audio,
                "processing_audio": processing_audio,
                "retrying_audio": retrying_audio,
                "success_rate": transcription_success_rate,
                "avg_confidence": avg_confidence,
                "top_languages": top_languages,
                "avg_processing_time_ms": avg_processing_time,
                "avg_audio_duration": avg_audio_length,
                "keywords": keywords_list
            }
        }, status=status.HTTP_200_OK)


class DashboardAnalyticsView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        print("=" * 60)
        print("DashboardAnalyticsView called")
        print("Authenticated:", request.user.is_authenticated)
        print("Username:", request.user)
        print("Staff:", request.user.is_staff)
        print("Superuser:", request.user.is_superuser)
        print("=" * 60)
        range_type = request.query_params.get('range_type')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Segment query filters
        category = request.query_params.get('category')
        status_param = request.query_params.get('status')
        project = request.query_params.get('project')
        location = request.query_params.get('location')

        complaints = Complaint.objects.all()

        # Apply segment filters
        if category:
            complaints = complaints.filter(category=category)
        if status_param:
            complaints = complaints.filter(status=status_param)
        if project:
            complaints = complaints.filter(project_id=project)
        if location:
            complaints = complaints.filter(location_id=location)

        # Apply date filters
        if range_type == 'today':
            today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
            complaints = complaints.filter(created_at__gte=today_start)
        elif range_type == '7days':
            seven_days_ago = timezone.now() - timedelta(days=7)
            complaints = complaints.filter(created_at__gte=seven_days_ago)
        elif range_type == '30days':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            complaints = complaints.filter(created_at__gte=thirty_days_ago)
        elif range_type == 'custom' or (not range_type and (start_date or end_date)):
            if start_date:
                complaints = complaints.filter(created_at__date__gte=start_date)
            if end_date:
                complaints = complaints.filter(created_at__date__lte=end_date)

        total_count = complaints.count()

        # Category breakdowns
        cat_map = defaultdict(lambda: {"Pending": 0, "In Progress": 0, "Completed": 0, "Resolved": 0, "Rejected": 0, "count": 0})
        for item in complaints.values('category', 'status').annotate(count=Count('id')):
            cat = item['category']
            st = item['status']
            cnt = item['count']
            cat_map[cat][st] = cnt
            cat_map[cat]["count"] += cnt

        category_data = []
        for cat, stats in cat_map.items():
            pct = round((stats["count"] / total_count) * 100, 1) if total_count > 0 else 0.0
            category_data.append({"category": cat, "percentage": pct, **stats})
        category_data = sorted(category_data, key=lambda x: x['count'], reverse=True)

        # Status counts
        status_map = defaultdict(int)
        for item in complaints.values('status').annotate(count=Count('id')):
            status_map[item['status']] = item['count']

        status_data = []
        for s, count in status_map.items():
            pct = round((count / total_count) * 100, 1) if total_count > 0 else 0.0
            status_data.append({"status": s, "count": count, "percentage": pct})
        status_data = sorted(status_data, key=lambda x: x['count'], reverse=True)

        # Projects breakdowns
        proj_map = defaultdict(lambda: {"Pending": 0, "In Progress": 0, "Completed": 0, "Resolved": 0, "Rejected": 0, "count": 0})
        for item in complaints.values('project__name', 'status').annotate(count=Count('id')):
            proj = item['project__name'] or 'Unknown Project'
            st = item['status']
            cnt = item['count']
            proj_map[proj][st] = cnt
            proj_map[proj]["count"] += cnt

        project_data = []
        for proj, stats in proj_map.items():
            pct = round((stats["count"] / total_count) * 100, 1) if total_count > 0 else 0.0
            project_data.append({"project": proj, "percentage": pct, **stats})
        project_data = sorted(project_data, key=lambda x: x['count'], reverse=True)

        # Locations breakdowns
        loc_map = defaultdict(lambda: {"Pending": 0, "In Progress": 0, "Completed": 0, "Resolved": 0, "Rejected": 0, "count": 0})
        for item in complaints.values('location__name', 'status').annotate(count=Count('id')):
            loc = item['location__name'] or 'Unknown Location'
            st = item['status']
            cnt = item['count']
            loc_map[loc][st] = cnt
            loc_map[loc]["count"] += cnt

        location_data = []
        for loc, stats in loc_map.items():
            pct = round((stats["count"] / total_count) * 100, 1) if total_count > 0 else 0.0
            location_data.append({"location": loc, "percentage": pct, **stats})
        location_data = sorted(location_data, key=lambda x: x['count'], reverse=True)

        # Monthly Trend breakdowns
        trend_map = defaultdict(
            lambda: {
                "Pending": 0,
                "In Progress": 0,
                "Completed": 0,
                "Resolved": 0,
                "Rejected": 0,
                "count": 0,
            }
        )

        monthly_data = (
            complaints
            .annotate(month=TruncMonth("created_at"))
            .values("month", "status")
            .annotate(count=Count("id"))
        )

        for item in monthly_data:
            month = item["month"].strftime("%Y-%m")
            complaint_status = item["status"]
            count = item["count"]

            trend_map[month][complaint_status] = count
            trend_map[month]["count"] += count

        trend_data = []

        for month, stats in sorted(trend_map.items()):
            pct = round((stats["count"] / total_count) * 100, 1) if total_count > 0 else 0.0

            trend_data.append({
                "month": month,
                "percentage": pct,
                **stats
            })
        
        # Calculate drill-down list
        drilldown_complaints = ComplaintSerializer(complaints.order_by('-created_at'), many=True).data

        # Calculate summary statistics
        pending_c = status_map.get("Pending", 0)
        comp_c = status_map.get("Completed", 0)
        res_c = status_map.get("Resolved", 0)
        resolution_rate = round(((comp_c + res_c) / total_count * 100), 1) if total_count > 0 else 0.0
        pending_rate = round((pending_c / total_count * 100), 1) if total_count > 0 else 0.0

        summary_stats = {
            "total_complaints": total_count,
            "pending_count": pending_c,
            "in_progress_count": status_map.get("In Progress", 0),
            "completed_count": comp_c,
            "resolved_count": res_c,
            "rejected_count": status_map.get("Rejected", 0),
            "resolution_rate": resolution_rate,
            "pending_rate": pending_rate
        }

        return Response({
            "category_distribution": category_data,
            "status_distribution": status_data,
            "project_distribution": project_data,
            "location_distribution": location_data,
            "monthly_trend": trend_data,
            "drilldown_complaints": drilldown_complaints,
            "summary_stats": summary_stats
        }, status=status.HTTP_200_OK)


# --- Excel Data Export ---

class DashboardExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        queryset = Complaint.objects.all()

        project = request.query_params.get('project')
        if project:
            queryset = queryset.filter(project_id=project)

        location = request.query_params.get('location')
        if location:
            queryset = queryset.filter(location_id=location)

        category = request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)

        business_unit = request.query_params.get('business_unit')
        if business_unit:
            queryset = queryset.filter(business_unit__iexact=business_unit)

        status = request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)

        language = request.query_params.get('language')
        if language:
            queryset = queryset.filter(language=language)

        range_type = request.query_params.get('range_type')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if range_type == 'today':
            today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
            queryset = queryset.filter(created_at__gte=today_start)
        elif range_type == '7days':
            seven_days_ago = timezone.now() - timedelta(days=7)
            queryset = queryset.filter(created_at__gte=seven_days_ago)
        elif range_type == '30days':
            thirty_days_ago = timezone.now() - timedelta(days=30)
            queryset = queryset.filter(created_at__gte=thirty_days_ago)
        elif range_type == 'custom' or (not range_type and (start_date or end_date)):
            if start_date:
                queryset = queryset.filter(created_at__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(created_at__date__lte=end_date)

        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(reference_number__icontains=search) |
                models.Q(project__name__icontains=search) |
                models.Q(location__name__icontains=search)
            )

        queryset = queryset.order_by('-created_at')

        count = queryset.count()
        if count > 10000:
            return Response(
                {"detail": "Export exceeds maximum limit of 10,000 records. Please apply additional filters."},
                status=status.HTTP_400_BAD_REQUEST
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Complaints"

        headers = [
            "Reference Number",
            "Project",
            "Business Unit",
            "Location",
            "Category",
            "Language",
            "Status",
            "Created At",
            "Updated At",
            "Photo URL",
            "Audio URL",
            "Original Text",
            "English Translation",
            
            # Phase 5.6 fields
            "Worker Selected Language",
            "Detected Language",
            "Transcript Confidence",
            "Translation Confidence",
            "Speech Processing Duration (ms)",
            "Translation Language Pair",
            "Translation Verification Result"
        ]

        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.row_dimensions[1].height = 28

        for item in queryset:
            created_str = item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else ''
            updated_str = item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if item.updated_at else ''
            ws.append([
                item.reference_number or '',
                get_bilingual_name(item.project, item.language),
                item.business_unit or '',
                get_bilingual_name(item.location, item.language),
                get_bilingual_category(item.category, item.language),
                item.language or '',
                item.status or '',
                created_str,
                updated_str,
                item.photo_url or '',
                item.audio_url or '',
                item.original_text or '',
                item.english_translation or '',
                
                # Phase 5.6 fields
                item.worker_selected_language or '',
                item.detected_language or '',
                item.transcript_confidence if item.transcript_confidence is not None else '',
                item.translation_confidence if item.translation_confidence is not None else '',
                item.speech_processing_duration_ms if item.speech_processing_duration_ms is not None else '',
                item.translation_language_pair or '',
                item.translation_verification_result or ''
            ])

        cell_font = Font(name="Segoe UI", size=10)
        cell_align = Alignment(vertical="center")

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.alignment = cell_align

        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        import io
        from django.http import HttpResponse
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="complaints_export.xlsx"'
        return response


# --- Notifications Views ---

class NotificationListView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        # Slice to the latest 100 notifications for production performance optimization
        notifications = Notification.objects.all()[:100]
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class NotificationReadView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request):
        notif_id = request.data.get('id')
        mark_all = request.data.get('all', False)
        now = timezone.now()
        
        if mark_all:
            Notification.objects.filter(is_read=False).update(is_read=True, read_at=now)
            return Response({"success": True, "detail": "All notifications marked as read."}, status=status.HTTP_200_OK)
            
        if notif_id:
            try:
                notification = Notification.objects.get(id=notif_id)
                notification.is_read = True
                notification.read_at = now
                notification.save()
                return Response(NotificationSerializer(notification).data, status=status.HTTP_200_OK)
            except Notification.DoesNotExist:
                return Response({"detail": "Notification not found."}, status=status.HTTP_404_NOT_FOUND)
                
        return Response({"detail": "Invalid parameters."}, status=status.HTTP_400_BAD_REQUEST)


# --- User Management Views (Strict Backend Superuser Restriction) ---

class UserListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Only Superusers can access User Management."}, status=status.HTTP_403_FORBIDDEN)
            
        User = get_user_model()
        users = User.objects.filter(is_staff=True).order_by('-date_joined')
        serializer = AdminUserSerializer(users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class UserCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Only Superusers can access User Management."}, status=status.HTTP_403_FORBIDDEN)
            
        username = request.data.get('username')
        email = request.data.get('email')
        password = request.data.get('password')
        confirm_password = request.data.get('confirm_password')
        role = request.data.get('role') # 'Super Admin' or 'Admin'
        
        if not username or not password or not role:
            return Response({"detail": "Please fill in all required fields."}, status=status.HTTP_400_BAD_REQUEST)
            
        if password != confirm_password:
            return Response({"detail": "Passwords do not match."}, status=status.HTTP_400_BAD_REQUEST)
            
        User = get_user_model()
        if User.objects.filter(username=username).exists():
            return Response({"detail": "Username is already taken."}, status=status.HTTP_400_BAD_REQUEST)
            
        is_superuser = (role == 'Super Admin')
        
        user = User.objects.create_user(
            username=username,
            email=email or "",
            password=password,
            is_staff=True,
            is_superuser=is_superuser
        )
        
        # Trigger Admin Created Notification
        Notification.objects.create(
            title="New Admin Created",
            description=f"Administrator '{username}' ({role}) has been created by {request.user.username}.",
            notification_type="user_created"
        )
        
        return Response(AdminUserSerializer(user).data, status=status.HTTP_201_CREATED)


class UserPasswordResetView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Only Superusers can access User Management."}, status=status.HTTP_403_FORBIDDEN)
            
        user_id = request.data.get('user_id')
        new_password = request.data.get('new_password')
        
        if not user_id or not new_password:
            return Response({"detail": "Missing parameters."}, status=status.HTTP_400_BAD_REQUEST)
            
        User = get_user_model()
        try:
            user = User.objects.get(id=user_id)
            
            # Validate password strength using django validators (unless in testing to allow simple passwords)
            if not getattr(settings, 'TESTING', False):
                from django.contrib.auth.password_validation import validate_password
                from django.core.exceptions import ValidationError
                try:
                    validate_password(new_password, user=user)
                except ValidationError as e:
                    return Response({"detail": e.messages[0]}, status=status.HTTP_400_BAD_REQUEST)
                
            user.set_password(new_password)
            user.save()
            
            # Invalidate all active sessions for the user to force immediate re-authentication
            from django.contrib.sessions.models import Session
            from django.utils import timezone
            for session in Session.objects.filter(expire_date__gte=timezone.now()):
                try:
                    data = session.get_decoded()
                    if data.get('_auth_user_id') == str(user.id) or data.get('_auth_user_id') == user.id:
                        session.delete()
                except Exception:
                    pass
                    
            return Response({"success": True, "detail": "Password reset successfully."}, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)


class UserToggleActiveView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Only Superusers can access User Management."}, status=status.HTTP_403_FORBIDDEN)
            
        user_id = request.data.get('user_id')
        is_active = request.data.get('is_active')
        
        if user_id is None or is_active is None:
            return Response({"detail": "Missing parameters."}, status=status.HTTP_400_BAD_REQUEST)
            
        # Lockout Protection Check (Self-Disabling Prevention)
        if int(user_id) == request.user.id and not is_active:
            return Response(
                {"detail": "You cannot disable your own active account."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        User = get_user_model()
        try:
            user = User.objects.get(id=user_id)
            user.is_active = is_active
            
            # Allow editing email and role (is_superuser) at the same time if requested
            email = request.data.get('email')
            role = request.data.get('role')
            if email is not None:
                user.email = email
            if role is not None:
                user.is_superuser = (role == 'Super Admin')
                
            user.save()
            
            if not is_active:
                # Trigger User Disabled Notification
                Notification.objects.create(
                    title="Administrator Disabled",
                    description=f"Admin account '{user.username}' has been disabled by {request.user.username}.",
                    notification_type="user_disabled"
                )
                
            return Response(AdminUserSerializer(user).data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)


# --- Audit Logs Views ---

class ActivityLogView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        activities = UserActivity.objects.all()[:200]  # limit to latest 200 logs
        serializer = UserActivitySerializer(activities, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class StatusHistoryView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        complaint_id = request.query_params.get('complaint_id')
        queryset = ComplaintStatusHistory.objects.all()
        if complaint_id:
            queryset = queryset.filter(complaint_id=complaint_id)
            
        serializer = ComplaintStatusHistorySerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


# --- Master Data Management Views ---

class MasterProjectView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        projects = Project.objects.all().order_by('name')
        return Response(ProjectSerializer(projects, many=True).data)

    def post(self, request):
        serializer = ProjectSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class MasterProjectDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def put(self, request, pk):
        try:
            project = Project.objects.get(pk=pk)
        except Project.DoesNotExist:
            return Response({"detail": "Project not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = ProjectSerializer(project, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        try:
            project = Project.objects.get(pk=pk)
        except Project.DoesNotExist:
            return Response({"detail": "Project not found."}, status=status.HTTP_404_NOT_FOUND)
        if Complaint.objects.filter(project=project).exists():
            project.is_active = False
            project.save(update_fields=['is_active'])
            return Response({"detail": "Project deactivated (has linked complaints)."}, status=status.HTTP_200_OK)
        project.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MasterLocationView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        project_id = request.query_params.get('project')
        qs = Location.objects.select_related('project').all().order_by('project__name', 'name')
        if project_id:
            qs = qs.filter(project_id=project_id)
        return Response(LocationSerializer(qs, many=True).data)

    def post(self, request):
        serializer = LocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class MasterLocationDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def put(self, request, pk):
        try:
            location = Location.objects.get(pk=pk)
        except Location.DoesNotExist:
            return Response({"detail": "Location not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = LocationSerializer(location, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        try:
            location = Location.objects.get(pk=pk)
        except Location.DoesNotExist:
            return Response({"detail": "Location not found."}, status=status.HTTP_404_NOT_FOUND)
        if Complaint.objects.filter(location=location).exists():
            location.is_active = False
            location.save(update_fields=['is_active'])
            return Response({"detail": "Location deactivated (has linked complaints)."}, status=status.HTTP_200_OK)
        location.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MasterCategoryView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        categories = Category.objects.all().order_by('sort_order', 'label')
        return Response(CategorySerializer(categories, many=True).data)

    def post(self, request):
        serializer = CategorySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class MasterCategoryDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def put(self, request, pk):
        try:
            category = Category.objects.get(pk=pk)
        except Category.DoesNotExist:
            return Response({"detail": "Category not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = CategorySerializer(category, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        try:
            category = Category.objects.get(pk=pk)
        except Category.DoesNotExist:
            return Response({"detail": "Category not found."}, status=status.HTTP_404_NOT_FOUND)
        category.is_active = False
        category.save(update_fields=['is_active'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class MasterLanguageView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        languages = Language.objects.all().order_by('sort_order', 'name')
        return Response(LanguageSerializer(languages, many=True).data)

    def post(self, request):
        serializer = LanguageSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class MasterLanguageDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def put(self, request, pk):
        try:
            language = Language.objects.get(pk=pk)
        except Language.DoesNotExist:
            return Response({"detail": "Language not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = LanguageSerializer(language, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        try:
            language = Language.objects.get(pk=pk)
        except Language.DoesNotExist:
            return Response({"detail": "Language not found."}, status=status.HTTP_404_NOT_FOUND)
        language.is_active = False
        language.save(update_fields=['is_active'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class MasterBusinessUnitView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        units = BusinessUnit.objects.all().order_by('name')
        return Response([{"id": u.id, "name": u.name, "is_active": u.is_active} for u in units])

    def post(self, request):
        name = (request.data.get('name') or '').strip()
        is_active = request.data.get('is_active', True)
        if not name:
            return Response({"detail": "Business unit name is required."}, status=status.HTTP_400_BAD_REQUEST)
        if BusinessUnit.objects.filter(name__iexact=name).exists():
            return Response({"detail": "Business unit with this name already exists."}, status=status.HTTP_400_BAD_REQUEST)
        bu = BusinessUnit.objects.create(name=name, is_active=is_active)
        return Response({"id": bu.id, "name": bu.name, "is_active": bu.is_active}, status=status.HTTP_201_CREATED)


class MasterBusinessUnitDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def put(self, request, pk):
        try:
            bu = BusinessUnit.objects.get(pk=pk)
        except BusinessUnit.DoesNotExist:
            return Response({"detail": "Business unit not found."}, status=status.HTTP_404_NOT_FOUND)
        
        name = (request.data.get('name') or '').strip()
        is_active = request.data.get('is_active')
        
        if name:
            if BusinessUnit.objects.filter(name__iexact=name).exclude(pk=pk).exists():
                return Response({"detail": "Business unit with this name already exists."}, status=status.HTTP_400_BAD_REQUEST)
            old_name = bu.name
            bu.name = name
            Project.objects.filter(business_unit=old_name).update(business_unit=name)
            Complaint.objects.filter(business_unit=old_name).update(business_unit=name)
            
        if is_active is not None:
            bu.is_active = is_active
            
        bu.save()
        return Response({"id": bu.id, "name": bu.name, "is_active": bu.is_active})

    def delete(self, request, pk):
        try:
            bu = BusinessUnit.objects.get(pk=pk)
        except BusinessUnit.DoesNotExist:
            return Response({"detail": "Business unit not found."}, status=status.HTTP_404_NOT_FOUND)
            
        if Project.objects.filter(business_unit=bu.name).exists():
            bu.is_active = False
            bu.save()
            return Response({"detail": "Business unit deactivated (has linked projects)."}, status=status.HTTP_200_OK)
        bu.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MasterSettingView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        settings = SystemSetting.objects.all().order_by('key')
        return Response([{"id": s.id, "key": s.key, "value": s.value, "description": s.description} for s in settings])

    def post(self, request):
        key = (request.data.get('key') or '').strip().upper()
        value = (request.data.get('value') or '').strip()
        description = (request.data.get('description') or '').strip()
        
        if not key:
            return Response({"detail": "Setting key is required."}, status=status.HTTP_400_BAD_REQUEST)
        if SystemSetting.objects.filter(key=key).exists():
            return Response({"detail": "Setting with this key already exists."}, status=status.HTTP_400_BAD_REQUEST)
            
        setting = SystemSetting.objects.create(key=key, value=value, description=description)
        return Response({"id": setting.id, "key": setting.key, "value": setting.value, "description": setting.description}, status=status.HTTP_201_CREATED)


class MasterSettingDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def put(self, request, pk):
        try:
            setting = SystemSetting.objects.get(pk=pk)
        except SystemSetting.DoesNotExist:
            return Response({"detail": "Setting not found."}, status=status.HTTP_404_NOT_FOUND)
            
        value = request.data.get('value')
        description = request.data.get('description')
        
        if value is not None:
            setting.value = str(value).strip()
        if description is not None:
            setting.description = str(description).strip()
            
        setting.save()
        return Response({"id": setting.id, "key": setting.key, "value": setting.value, "description": setting.description})

    def delete(self, request, pk):
        try:
            setting = SystemSetting.objects.get(pk=pk)
        except SystemSetting.DoesNotExist:
            return Response({"detail": "Setting not found."}, status=status.HTTP_404_NOT_FOUND)
        setting.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

